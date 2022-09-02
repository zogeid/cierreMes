import datetime
import glob
import os

from openpyxl import load_workbook

import Utils as U
from Pep import Pep

filled_peps = {}


# LEER LA INFORMACION DE LA SIMULACION DE OIAB
def leer_simulacion():
    wb = load_workbook('Simulación.xlsx')
    sheet = wb['Simulación']
    max_row = wb.active.max_row
    col = 'A'
    peps = []

    # recorre Simulacion.xlsx para obtener los diferentes PEPs y guardarlos en peps
    for row in range(2, max_row):
        cell_pep = sheet[col + row.__str__()].value
        if cell_pep is not None and cell_pep not in peps:
            peps.append(cell_pep)

    # Para cada PEP recoge datos de concepto, mob, nombre
    # En funcion del concepto acumula los importes de Ingreso, Coste y MOB
    # Con esta información genera un objeto PEP y lo añade a filled_peps
    for i in peps:
        ingreso_por_recurso_propio = 0
        coste_por_recurso_propio = 0
        mob = 0
        nombre_proyecto = ''

        for row in range(2, max_row+1):
            cell_pep = col + row.__str__()

            if sheet[cell_pep].value is not None and sheet[cell_pep].value == i:
                cell_concepto = sheet['AK' + row.__str__()].value
                cell_mob = sheet['AJ' + row.__str__()].value
                cell_nombre = sheet['B' + row.__str__()].value
                cell_importe = sheet['AR' + row.__str__()].value

                if cell_concepto is not None:
                    if 'INGRESO POR RECURSO PROPIO' in cell_concepto:
                        ingreso_por_recurso_propio = ingreso_por_recurso_propio + cell_importe
                    if 'COSTE POR RECURSO PROPIO' in cell_concepto:
                        coste_por_recurso_propio = coste_por_recurso_propio + cell_importe
                if cell_mob is not None:
                    if 'MOB-' in cell_mob:
                        mob = mob + cell_importe
                if cell_nombre is not None:
                    nombre_proyecto = cell_nombre

        p = Pep(i, ingreso_por_recurso_propio, coste_por_recurso_propio, mob, nombre_proyecto)
        filled_peps[p.pep] = p


# GRABAR INFORMACION EN LOS FRP'S
def grabar_frp():
    print(f'Grabando FRP de: {U.get_current_month()}\n')
    # recorrer los xlsx
    gen = (f for f in glob.glob(os.path.join('*.xlsx')) if 'FRP' in f)

    for filename in gen:
        pep = filename[filename.find('D-'):filename.find('D-')+13]
        current_p = filled_peps[pep]
        wb = load_workbook(filename)
        sheet = wb.active

        if U.get_current_year() == 2021:  # row2
            sheet['F2'].value = U.get_current_month()
        elif U.get_current_year() == 2022:  # row26
            sheet['F26'].value = U.get_current_month()

            suma_mano_obra = 0
            for i in range(29, 42): # IMPORTANT! Si se añaden más perfiles/tarifas nuevas comprobar el rango
                a = sheet[f'C{i}'].value
                b = sheet[f'{U.get_current_month_column()}{i}'].value

                if a is not None and b is not None:
                    try:
                        a = float(a)
                        b = float(b)
                        if a > 0 and b > 0:
                            suma_mano_obra = suma_mano_obra + (a * b)
                    except ValueError:
                        print(f'{current_p.nombre}: Revisa el formato de excel FRP\n')


            if round(suma_mano_obra, 0) == round(current_p.coste, 0):  # Comprobamos que costes coinciden
                sheet[f'{U.get_current_month_column()}47'].value = current_p.mob
                sheet[f'{U.get_current_month_column()}52'].value = current_p.ingreso

                try:
                    costes = suma_mano_obra + current_p.mob
                    ingreso_mob = costes / (1-sheet[f'{U.get_current_month_column()}50'].value)
                    reg = round(ingreso_mob - current_p.ingreso, 2)
                except TypeError as e:
                    print(f'Error calculando regularizacion, revisa los datos de {pep}{current_p.nombre}')
                    break

                filled_peps[pep].cierre_coste = costes
                filled_peps[pep].cierre_ingreso = ingreso_mob
                filled_peps[pep].regularizacion = reg
                filled_peps[pep].__repr__()
                wb.save(filename)

            else:
                print(f'{pep} {current_p.nombre}: Costes_mano_obra do not match')


def generar_informe_cierre():
    filename = f'{U.get_current_year()}{datetime.datetime.now().month} Cierre.xlsx'
    wb = load_workbook(filename)
    sheet = wb.active

    for c in range(4, 10):
        try:
            pep = sheet[f'B{c}'].value
            sheet[f'D{c}'].value = filled_peps[pep].cierre_coste
            sheet[f'F{c}'].value = filled_peps[pep].cierre_ingreso
            sheet[f'G{c}'].value = filled_peps[pep].regularizacion
        except KeyError as e:
            print(f'Iteration {c} KeyError exception {e}. Under control!')
    wb.save(filename)


leer_simulacion()
for r in filled_peps: r.__repr__()
grabar_frp()
generar_informe_cierre()
